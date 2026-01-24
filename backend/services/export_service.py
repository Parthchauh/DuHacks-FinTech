"""
Export Service - Portfolio Export to Excel and PDF
===================================================
Generates downloadable Excel spreadsheets and PDF reports
for user portfolios with holdings, transactions, and metrics.
"""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


def generate_excel(
    portfolio_name: str,
    holdings: List[Dict[str, Any]],
    transactions: List[Dict[str, Any]],
    metrics: Dict[str, Any]
) -> bytes:
    """
    Generate an Excel workbook with portfolio data.
    
    Returns bytes that can be sent as a file download.
    """
    wb = Workbook()
    
    # ===== Summary Sheet =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    # Title
    ws_summary["A1"] = f"OptiWealth Portfolio Report: {portfolio_name}"
    ws_summary["A1"].font = Font(bold=True, size=18)
    ws_summary.merge_cells("A1:D1")
    
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary["A2"].font = Font(italic=True, size=10)
    
    # Metrics section
    row = 4
    metrics_data = [
        ("Total Value", f"₹{metrics.get('total_value', 0):,.2f}"),
        ("Total Invested", f"₹{metrics.get('total_invested', 0):,.2f}"),
        ("Total Return", f"₹{metrics.get('total_return', 0):,.2f}"),
        ("Return %", f"{metrics.get('total_return_percent', 0):.2f}%"),
        ("Sharpe Ratio", f"{metrics.get('sharpe_ratio', 0):.2f}"),
        ("Volatility", f"{metrics.get('volatility', 0):.2f}%"),
        ("Beta", f"{metrics.get('beta', 0):.2f}"),
        ("Risk Score", f"{metrics.get('risk_score', 0)}/10"),
    ]
    
    for label, value in metrics_data:
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)
        row += 1
    
    # Adjust column widths
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 20
    
    # ===== Holdings Sheet =====
    ws_holdings = wb.create_sheet("Holdings")
    
    holdings_headers = ["Ticker", "Name", "Type", "Quantity", "Avg Price", "Current Price", 
                       "Current Value", "P&L", "P&L %", "Allocation %"]
    
    for col, header in enumerate(holdings_headers, 1):
        cell = ws_holdings.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, holding in enumerate(holdings, 2):
        ws_holdings.cell(row=row_idx, column=1, value=holding.get("ticker", ""))
        ws_holdings.cell(row=row_idx, column=2, value=holding.get("name", ""))
        ws_holdings.cell(row=row_idx, column=3, value=holding.get("asset_type", "EQUITY"))
        ws_holdings.cell(row=row_idx, column=4, value=holding.get("quantity", 0))
        ws_holdings.cell(row=row_idx, column=5, value=holding.get("avg_buy_price", 0))
        ws_holdings.cell(row=row_idx, column=6, value=holding.get("current_price", 0))
        ws_holdings.cell(row=row_idx, column=7, value=holding.get("current_value", 0))
        ws_holdings.cell(row=row_idx, column=8, value=holding.get("profit_loss", 0))
        ws_holdings.cell(row=row_idx, column=9, value=holding.get("profit_loss_percent", 0))
        ws_holdings.cell(row=row_idx, column=10, value=holding.get("actual_allocation", 0))
    
    # Auto-adjust column widths
    for col in range(1, len(holdings_headers) + 1):
        ws_holdings.column_dimensions[get_column_letter(col)].width = 15
    
    # ===== Transactions Sheet =====
    ws_transactions = wb.create_sheet("Transactions")
    
    txn_headers = ["Date", "Ticker", "Type", "Quantity", "Price", "Total Amount", "Notes"]
    
    for col, header in enumerate(txn_headers, 1):
        cell = ws_transactions.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, txn in enumerate(transactions, 2):
        executed_at = txn.get("executed_at", "")
        if isinstance(executed_at, datetime):
            executed_at = executed_at.strftime("%Y-%m-%d")
        ws_transactions.cell(row=row_idx, column=1, value=executed_at)
        ws_transactions.cell(row=row_idx, column=2, value=txn.get("ticker", ""))
        ws_transactions.cell(row=row_idx, column=3, value=txn.get("transaction_type", ""))
        ws_transactions.cell(row=row_idx, column=4, value=txn.get("quantity", 0))
        ws_transactions.cell(row=row_idx, column=5, value=txn.get("price", 0))
        ws_transactions.cell(row=row_idx, column=6, value=txn.get("total_amount", 0))
        ws_transactions.cell(row=row_idx, column=7, value=txn.get("notes", ""))
    
    # Auto-adjust column widths
    for col in range(1, len(txn_headers) + 1):
        ws_transactions.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_pdf(
    portfolio_name: str,
    holdings: List[Dict[str, Any]],
    transactions: List[Dict[str, Any]],
    metrics: Dict[str, Any]
) -> bytes:
    """
    Generate a PDF report with portfolio data.
    
    Returns bytes that can be sent as a file download.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        alignment=TA_CENTER,
        spaceAfter=20,
        textColor=colors.HexColor("#3B82F6")
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.gray,
        spaceAfter=30
    )
    
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.HexColor("#1E293B")
    )
    
    # Title
    elements.append(Paragraph(f"OptiWealth Portfolio Report", title_style))
    elements.append(Paragraph(portfolio_name, styles['Heading2']))
    elements.append(Paragraph(
        f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M')}",
        subtitle_style
    ))
    
    # Portfolio Summary
    elements.append(Paragraph("Portfolio Summary", section_style))
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Value", f"₹{metrics.get('total_value', 0):,.2f}"],
        ["Total Invested", f"₹{metrics.get('total_invested', 0):,.2f}"],
        ["Total Return", f"₹{metrics.get('total_return', 0):,.2f}"],
        ["Return %", f"{metrics.get('total_return_percent', 0):.2f}%"],
        ["Sharpe Ratio", f"{metrics.get('sharpe_ratio', 0):.2f}"],
        ["Risk Score", f"{metrics.get('risk_score', 0)}/10 ({metrics.get('risk_level', 'Moderate')})"],
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3B82F6")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Holdings
    elements.append(Paragraph("Holdings", section_style))
    
    holdings_data = [["Ticker", "Name", "Qty", "Avg Price", "Value", "P&L %"]]
    for h in holdings[:20]:  # Limit to 20 for PDF
        holdings_data.append([
            h.get("ticker", ""),
            h.get("name", "")[:20],  # Truncate long names
            f"{h.get('quantity', 0):.2f}",
            f"₹{h.get('avg_buy_price', 0):,.0f}",
            f"₹{h.get('current_value', 0):,.0f}",
            f"{h.get('profit_loss_percent', 0):.1f}%"
        ])
    
    holdings_table = Table(holdings_data, colWidths=[0.8*inch, 1.5*inch, 0.6*inch, 0.9*inch, 1*inch, 0.7*inch])
    holdings_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3B82F6")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    elements.append(holdings_table)
    elements.append(Spacer(1, 20))
    
    # Recent Transactions
    if transactions:
        elements.append(Paragraph("Recent Transactions", section_style))
        
        txn_data = [["Date", "Ticker", "Type", "Qty", "Amount"]]
        for t in transactions[:10]:  # Last 10 transactions
            executed_at = t.get("executed_at", "")
            if isinstance(executed_at, datetime):
                executed_at = executed_at.strftime("%Y-%m-%d")
            elif isinstance(executed_at, str) and "T" in executed_at:
                executed_at = executed_at.split("T")[0]
            
            txn_data.append([
                executed_at,
                t.get("ticker", ""),
                t.get("transaction_type", ""),
                f"{t.get('quantity', 0):.2f}",
                f"₹{t.get('total_amount', 0):,.0f}"
            ])
        
        txn_table = Table(txn_data, colWidths=[1*inch, 1*inch, 0.8*inch, 0.8*inch, 1.2*inch])
        txn_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3B82F6")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        elements.append(txn_table)
    
    # Footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, 
                                   alignment=TA_CENTER, textColor=colors.gray)
    elements.append(Paragraph("Generated by OptiWealth - Your Smart Portfolio Manager", footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
